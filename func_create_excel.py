# Import libraries
import openpyxl
import numpy as np
import pandas as pd

file_name = '4_backtest_convert.xlsx'

def create_excel():
  # Read pandas Dataframe and convert to XLSX Writer object. 
  bt_file = pd.read_csv('3_backtest_file.csv')


  # Replace empty values in Zscore column with NaN for deletion
  bt_file['Zscore'].replace('', np.nan, inplace=True )

  # Drop rows with NaN values in Zscore and first column
  bt_file.dropna(subset=['Zscore'] , inplace = True)
  bt_file.drop(columns=bt_file.columns[0], inplace = True)

  # Creates new openpyxl object from corrected empty z-scores above and places
  # in new sheet names below
  bt_file_two = bt_file.copy()
  writer = pd.ExcelWriter('backtest_convert.xlsx', engine='openpyxl')
  bt_file.to_excel(writer, sheet_name='n_steps')
  bt_file_two.to_excel(writer, sheet_name='mean_reverting')
  writer.save()


 
  # Read prior saved empty workbook
  nb = openpyxl.load_workbook('backtest_convert.xlsx')
  
  # Reads backtest file which will replace first 5 columns
  ob = openpyxl.load_workbook('Backtest_Calculations.xlsx')
  active_sheet = nb.active
  last_row = active_sheet.max_row


  # Assigned variables for active sheet in prior workbook
  steps_sheet = nb['n_steps']
  mean_sheet = nb['mean_reverting']

  # Assign variables to sheets in backtest calculations workbook
  ob_steps_sheet = ob['n_steps']
  ob_mean_sheet = ob['mean_reverting']

  # Loop through first 5 columns to replace values with new pairs
  for x in range(1,last_row+1):
    for y in range(1,6):
      ob_steps_sheet.cell(row=x, column=y).value = steps_sheet.cell(row=x, column=y).value
      ob_mean_sheet.cell(row=x, column=y).value = mean_sheet.cell(row=x, column=y).value
  
  # Saves new version of original backtest file
  ob.save(file_name)
  return file_name





######## IGNORE CODE BELOW THESE LINES ############
# # Assign variables to values for testing
# z_score_thresh = 1.0
# trading_capital = 1000
# rebate = -0.00025
# slippage = 0.0001
# close_long_steps = 5
# close_short_steps = 5


# # Assign inputs to sheet, Zscore threshold, trading capital, rebate, slippage etc
# tb_steps_sheet['AE1'].value = z_score_thresh
# tb_steps_sheet['AE2'].value = trading_capital
# tb_steps_sheet['AE3'].value = rebate
# tb_steps_sheet['AE4'].value = slippage

# # Backtest Results data
# long_profit = tb_steps_sheet['AF9'].value
# short_profit = tb_steps_sheet['AF10'].value
# net_profit = tb_steps_sheet['AF11'].value
# roi = tb_steps_sheet['AF13'].value
# long_wr = tb_steps_sheet['AF15'].value
# short_wr = tb_steps_sheet['AF16'].value
# avg_wr = tb_steps_sheet['AF17'].value
 

# row_range = steps_sheet[1:last_row]
######## IGNORE CODE ABOVE ###############

# nb = openpyxl.load_workbook('Backtest_Convt_Three.xlsx', data_only=True)

# nb_steps_sheet = nb['n_steps']
# nb_mean_sheet = nb['mean_reverting']
 
# # Backtest Results data on newly saved sheet
# nb_long_profit = nb_steps_sheet['AF9'].value
# nb_short_profit = nb_steps_sheet['AF10'].value
# nb_net_profit = nb_steps_sheet['AF11'].value
# nb_roi = nb_steps_sheet['AF13'].value
# nb_long_wr = nb_steps_sheet['AF15'].value
# nb_short_wr = nb_steps_sheet['AF16'].value
# nb_avg_wr = nb_steps_sheet['AF17'].value

# print(nb_net_profit)

##### IGNORE CODE ABOVE ############